<<<<<<< HEAD
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === CONFIG ===
folder_path = r"C:\operations"
output_file = os.path.join(folder_path, "processed_report.xlsx")

# === 1. Get latest CSV ===
files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
if not files:
    raise Exception("No CSV files found")

latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(folder_path, x)))
file_path = os.path.join(folder_path, latest_file)
print(f"Processing: {latest_file}")

# === 2. Load data ===
df = pd.read_csv(file_path)
df.columns = df.columns.str.strip().str.lower()

# === 3. Validate columns ===
required_cols = ['serialno', 'status', 'service', 'country', 'lcbalance', 'lctotal', 'date']
missing = [col for col in required_cols if col not in df.columns]
if missing:
    raise Exception(f"Missing columns: {missing}")

# === 4. Data prep ===
df['date'] = pd.to_datetime(df['date'], errors='coerce')
df['transaction amount'] = pd.to_numeric(df['lctotal'], errors='coerce')

# === 5. Status groups ===
exceptions_status = [
    'AUTHORIZED AND COMPLIANCE FAILED',
    'AUTHORIZED AND CORRESPONDENT FAILED',
    'AUTHORIZED AND PAYMENT HOLD',
    'CREATED'
]

intended_status = [
    'AUTHORIZED',
    'RETURNED',
    'RETURN REQUESTED',
    'REFUNDED'
]

# === 6. Exceptions ===
exceptions_df = df[df['status'].isin(exceptions_status)].copy()
now = datetime.now()

exceptions_df = exceptions_df[
    (exceptions_df['status'] != 'CREATED') |
    ((exceptions_df['status'] == 'CREATED') & (exceptions_df['date'] <= now - timedelta(hours=48)))
]

exceptions_df['short paid'] = None
mask = exceptions_df['status'] == 'AUTHORIZED AND PAYMENT HOLD'
exceptions_df.loc[mask, 'short paid'] = exceptions_df['lcbalance']

# === 7. Intended ===
intended_df = df[df['status'].isin(intended_status)].copy()

# === 8. Final columns ===
exceptions_cols = ['serialno', 'status', 'service', 'country', 'date', 'transaction amount', 'short paid']
intended_cols = ['serialno', 'status', 'service', 'country', 'date', 'transaction amount']

exceptions_output = exceptions_df[exceptions_cols]
intended_output = intended_df[intended_cols]

# === 9. Load workbook ===
if not os.path.exists(output_file):
    raise Exception("processed_report.xlsx not found.")

wb = load_workbook(output_file)

# === 9B. Ensure sheets exist ===
if 'Exception_Log' not in wb.sheetnames:
    ws_log = wb.create_sheet('Exception_Log')
    ws_log.append(exceptions_cols)
else:
    ws_log = wb['Exception_Log']

if 'Exception_Summary' not in wb.sheetnames:
    ws_summary = wb.create_sheet('Exception_Summary')
else:
    ws_summary = wb['Exception_Summary']

# === HELPER: Write data ===
def write_data(ws, df):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

# === HELPER: Ensure table ===
def ensure_table(ws, table_name):
    max_row = ws.max_row if ws.max_row > 1 else 2
    max_col = ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if ws.tables:
        table = list(ws.tables.values())[0]
        table.ref = ref
    else:
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

# === 10. Write main sheets ===
ws_ex = wb['Exceptions']
write_data(ws_ex, exceptions_output)
ensure_table(ws_ex, "ExceptionsTable")

ws_in = wb['Intended']
write_data(ws_in, intended_output)
ensure_table(ws_in, "IntendedTable")

# === 11. Append to Exception_Log (NO duplicates by serialno) ===
existing_serials = set()

for row in ws_log.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_serials.add(row[0])

for _, row in exceptions_output.iterrows():
    if row['serialno'] not in existing_serials:
        ws_log.append(row.tolist())

# === 12. Build Exception_Summary (COUNT PER STATUS ONLY) ===
log_data = list(ws_log.values)
headers = log_data[0]
data = log_data[1:]

log_df = pd.DataFrame(data, columns=headers)

summary_df = log_df.groupby('status').agg(
    exception_count=('status', 'count')
).reset_index()

summary_df = summary_df.sort_values(by='exception_count', ascending=False)

# Clear summary sheet
if ws_summary.max_row > 1:
    ws_summary.delete_rows(2, ws_summary.max_row)

# Write headers
for col_idx, col in enumerate(summary_df.columns, start=1):
    ws_summary.cell(row=1, column=col_idx, value=col)

# Write data
for r_idx, row in enumerate(summary_df.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws_summary.cell(row=r_idx, column=c_idx, value=value)

# === 13. Save ===
wb.save(output_file)

=======
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === CONFIG ===
folder_path = r"C:\operations"
output_file = os.path.join(folder_path, "processed_report.xlsx")

# === 1. Get latest CSV ===
files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
if not files:
    raise Exception("No CSV files found")

latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(folder_path, x)))
file_path = os.path.join(folder_path, latest_file)
print(f"Processing: {latest_file}")

# === 2. Load data ===
df = pd.read_csv(file_path)
df.columns = df.columns.str.strip().str.lower()

# === 3. Validate columns ===
required_cols = ['serialno', 'status', 'service', 'country', 'lcbalance', 'lctotal', 'date']
missing = [col for col in required_cols if col not in df.columns]
if missing:
    raise Exception(f"Missing columns: {missing}")

# === 4. Data prep ===
df['date'] = pd.to_datetime(df['date'], errors='coerce')
df['transaction amount'] = pd.to_numeric(df['lctotal'], errors='coerce')

# === 5. Status groups ===
exceptions_status = [
    'AUTHORIZED AND COMPLIANCE FAILED',
    'AUTHORIZED AND CORRESPONDENT FAILED',
    'AUTHORIZED AND PAYMENT HOLD',
    'CREATED'
]

intended_status = [
    'AUTHORIZED',
    'RETURNED',
    'RETURN REQUESTED',
    'REFUNDED'
]

# === 6. Exceptions ===
exceptions_df = df[df['status'].isin(exceptions_status)].copy()
now = datetime.now()

exceptions_df = exceptions_df[
    (exceptions_df['status'] != 'CREATED') |
    ((exceptions_df['status'] == 'CREATED') & (exceptions_df['date'] <= now - timedelta(hours=48)))
]

exceptions_df['short paid'] = None
mask = exceptions_df['status'] == 'AUTHORIZED AND PAYMENT HOLD'
exceptions_df.loc[mask, 'short paid'] = exceptions_df['lcbalance']

# === 7. Intended ===
intended_df = df[df['status'].isin(intended_status)].copy()

# === 8. Final columns ===
exceptions_cols = ['serialno', 'status', 'service', 'country', 'date', 'transaction amount', 'short paid']
intended_cols = ['serialno', 'status', 'service', 'country', 'date', 'transaction amount']

exceptions_output = exceptions_df[exceptions_cols]
intended_output = intended_df[intended_cols]

# === 9. Load workbook ===
if not os.path.exists(output_file):
    raise Exception("processed_report.xlsx not found.")

wb = load_workbook(output_file)

# === 9B. Ensure sheets exist ===
if 'Exception_Log' not in wb.sheetnames:
    ws_log = wb.create_sheet('Exception_Log')
    ws_log.append(exceptions_cols)
else:
    ws_log = wb['Exception_Log']

if 'Exception_Summary' not in wb.sheetnames:
    ws_summary = wb.create_sheet('Exception_Summary')
else:
    ws_summary = wb['Exception_Summary']

# === HELPER: Write data ===
def write_data(ws, df):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

# === HELPER: Ensure table ===
def ensure_table(ws, table_name):
    max_row = ws.max_row if ws.max_row > 1 else 2
    max_col = ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if ws.tables:
        table = list(ws.tables.values())[0]
        table.ref = ref
    else:
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

# === 10. Write main sheets ===
ws_ex = wb['Exceptions']
write_data(ws_ex, exceptions_output)
ensure_table(ws_ex, "ExceptionsTable")

ws_in = wb['Intended']
write_data(ws_in, intended_output)
ensure_table(ws_in, "IntendedTable")

# === 11. Append to Exception_Log (NO duplicates by serialno) ===
existing_serials = set()

for row in ws_log.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_serials.add(row[0])

for _, row in exceptions_output.iterrows():
    if row['serialno'] not in existing_serials:
        ws_log.append(row.tolist())

# === 12. Build Exception_Summary (COUNT PER STATUS ONLY) ===
log_data = list(ws_log.values)
headers = log_data[0]
data = log_data[1:]

log_df = pd.DataFrame(data, columns=headers)

summary_df = log_df.groupby('status').agg(
    exception_count=('status', 'count')
).reset_index()

summary_df = summary_df.sort_values(by='exception_count', ascending=False)

# Clear summary sheet
if ws_summary.max_row > 1:
    ws_summary.delete_rows(2, ws_summary.max_row)

# Write headers
for col_idx, col in enumerate(summary_df.columns, start=1):
    ws_summary.cell(row=1, column=col_idx, value=col)

# Write data
for r_idx, row in enumerate(summary_df.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws_summary.cell(row=r_idx, column=c_idx, value=value)

# === 13. Save ===
wb.save(output_file)

>>>>>>> a6b99c31b08f4d85982dd44adc8799e1a8a27d79
print("✅ SUCCESS: Exception summary (count per category) built")